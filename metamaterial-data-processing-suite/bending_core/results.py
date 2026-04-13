from __future__ import annotations

from pathlib import Path
import re

import numpy as np

from shared_core.adaptive_fit import (
    DECISION_ACCEPTED_CLUSTERED,
    DECISION_REJECTED_NO_FIT,
    DECISION_REJECTED_OUTLIER,
    DIAGNOSTIC_HEADERS,
    build_subgroup_band,
    canonical_sample_name,
    format_sample_display_name,
    is_within_subgroup_band,
)


RESULT_HEADERS = [
    "Sample Name",
    "Young's Modulus (GPa)",
    "Yield Strength (MPa)",
    "Ultimate Strength (MPa)",
    "Graph",
    "Valid",
]
DETAIL_RESULT_HEADERS = RESULT_HEADERS + DIAGNOSTIC_HEADERS
DETAIL_RESULT_COLUMN_COUNT = len(DETAIL_RESULT_HEADERS)


def _is_missing(value) -> bool:
    return value is None or (isinstance(value, (float, np.floating)) and np.isnan(value))


def _is_present(value) -> bool:
    return not _is_missing(value)


def get_subgroup(sample_name):
    if _is_missing(sample_name):
        return ""
    return re.split(r"\s*-\s*\d+$", canonical_sample_name(sample_name))[0]


def _ensure_detail_headers(ws):
    for column, header in enumerate(DETAIL_RESULT_HEADERS, start=1):
        ws.cell(row=1, column=column, value=header)


def _detail_result_column_map(ws):
    _ensure_detail_headers(ws)
    headers = [ws.cell(row=1, column=c).value for c in range(1, DETAIL_RESULT_COLUMN_COUNT + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    missing = [header for header in DETAIL_RESULT_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Missing required result headers: {', '.join(missing)}")

    return {header: index + 1 for index, header in enumerate(headers)}


def build_bending_result_row(
    *,
    sample_name: str,
    E: float,
    yield_strength: float,
    ultimate_strength: float,
    plot_path: Path,
    is_valid: bool,
    fit_start_strain: float,
    fit_end_strain: float,
    fit_mode: str,
    decision_reason: str,
):
    display_name = format_sample_display_name(sample_name, is_valid, decision_reason)
    return [
        display_name,
        E / 1e3 if _is_present(E) else np.nan,
        yield_strength if _is_present(yield_strength) else np.nan,
        ultimate_strength if _is_present(ultimate_strength) else np.nan,
        f'=HYPERLINK("{plot_path}", "Open Graph")',
        "Yes" if is_valid else "No",
        fit_start_strain if _is_present(fit_start_strain) else np.nan,
        fit_end_strain if _is_present(fit_end_strain) else np.nan,
        fit_mode,
        decision_reason,
    ]


def write_bending_results(output_path: Path, rows):
    from openpyxl import Workbook

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.append(DETAIL_RESULT_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(output_path)


def save_results_to_xlsx(
    sample_name: str,
    E: float,
    yield_strength: float,
    ultimate_strength: float,
    output_path: Path,
    plot_path: Path,
    is_valid: bool,
    *,
    fit_start_strain: float = np.nan,
    fit_end_strain: float = np.nan,
    fit_mode: str = "",
    decision_reason: str = "",
):
    from openpyxl import Workbook, load_workbook

    row_values = build_bending_result_row(
        sample_name=sample_name,
        E=E,
        yield_strength=yield_strength,
        ultimate_strength=ultimate_strength,
        plot_path=plot_path,
        is_valid=is_valid,
        fit_start_strain=fit_start_strain,
        fit_end_strain=fit_end_strain,
        fit_mode=fit_mode,
        decision_reason=decision_reason,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb.active
        _ensure_detail_headers(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(DETAIL_RESULT_HEADERS)

    canonical_input = canonical_sample_name(sample_name)
    found_row = None
    for row in ws.iter_rows(min_row=2, max_col=1):
        if canonical_sample_name(row[0].value) == canonical_input:
            found_row = row[0].row
            break

    if found_row:
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=found_row, column=col, value=value)
    else:
        ws.append(row_values)

    wb.save(output_path)
    check_valid_column(output_path, get_subgroup(sample_name))


def check_valid_column(
    file_path,
    group_name,
    sample_col="Sample Name",
    E_col="Young's Modulus (GPa)",
    valid_col="Valid",
):
    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    ws = wb.active

    col_map = _detail_result_column_map(ws)
    sample_idx = col_map[sample_col]
    E_idx = col_map[E_col]
    valid_idx = col_map[valid_col]
    reason_idx = col_map["Decision Reason"]

    group_rows = []
    group_E = []

    for row in range(2, ws.max_row + 1):
        sample_name = ws.cell(row=row, column=sample_idx).value
        if get_subgroup(sample_name) == group_name:
            group_rows.append(row)
            val = ws.cell(row=row, column=E_idx).value
            try:
                group_E.append(float(val))
            except (TypeError, ValueError):
                group_E.append(np.nan)

    band = build_subgroup_band(group_E)

    for row, E in zip(group_rows, group_E):
        if not np.isfinite(E):
            ws.cell(row=row, column=valid_idx, value="No")
            ws.cell(row=row, column=reason_idx, value=DECISION_REJECTED_NO_FIT)
            sample_value = ws.cell(row=row, column=sample_idx).value
            ws.cell(
                row=row,
                column=sample_idx,
                value=format_sample_display_name(sample_value, False, DECISION_REJECTED_NO_FIT),
            )
        elif band is None or is_within_subgroup_band(E, band):
            ws.cell(row=row, column=valid_idx, value="Yes")
            if not ws.cell(row=row, column=reason_idx).value:
                ws.cell(row=row, column=reason_idx, value=DECISION_ACCEPTED_CLUSTERED)
            sample_value = ws.cell(row=row, column=sample_idx).value
            ws.cell(
                row=row,
                column=sample_idx,
                value=format_sample_display_name(sample_value, True, ws.cell(row=row, column=reason_idx).value),
            )
        else:
            ws.cell(row=row, column=valid_idx, value="No")
            ws.cell(row=row, column=reason_idx, value=DECISION_REJECTED_OUTLIER)
            sample_value = ws.cell(row=row, column=sample_idx).value
            ws.cell(
                row=row,
                column=sample_idx,
                value=format_sample_display_name(sample_value, False, DECISION_REJECTED_OUTLIER),
            )

    wb.save(file_path)
