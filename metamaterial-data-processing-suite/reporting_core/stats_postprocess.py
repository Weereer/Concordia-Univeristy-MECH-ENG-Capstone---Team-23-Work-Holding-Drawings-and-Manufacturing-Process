from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import math
import statistics
import re

from shared_core.adaptive_fit import canonical_sample_name

Summary_start_col = 12


def sort_results_workbook(filepath, sheet_name=None):
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    if ws.max_row <= 2:
        wb.save(filepath)
        return

    rows = []
    for r in range(2, ws.max_row + 1):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in row_data):
            continue
        rows.append(row_data)

    def sort_key(row):
        sample_name = str(row[0]).strip() if row[0] else ""

        match = re.match(r"^(.*?)(?:\s*-\s*(\d+))?$", sample_name)
        if match:
            base = match.group(1).strip()
            num = int(match.group(2)) if match.group(2) else 0
        else:
            base = sample_name
            num = 0

        return (base, num)

    rows.sort(key=sort_key)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    for r_idx, row_data in enumerate(rows, start=2):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(filepath)


def get_subgroup(sample_name: str) -> str:
    if sample_name is None:
        return ""
    return re.split(r"\s*-\s*\d+$", canonical_sample_name(sample_name))[0]


def safe_cv(values):
    vals = []
    for v in values:
        try:
            x = float(v)
            if math.isfinite(x):
                vals.append(x)
        except Exception:
            pass

    if len(vals) < 2:
        return None

    mean_v = statistics.mean(vals)
    if mean_v == 0:
        return None

    stdev_v = statistics.stdev(vals)
    return stdev_v / mean_v


def build_summary_table(filepath, sheet_name=None, include_fracture=False):
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    raw_rows = []
    row = 2
    while ws[f"A{row}"].value not in (None, ""):
        sample_name = ws[f"A{row}"].value
        E = ws[f"B{row}"].value
        Y = ws[f"C{row}"].value
        U = ws[f"D{row}"].value

        if include_fracture:
            F = ws[f"E{row}"].value
            valid = ws[f"G{row}"].value
        else:
            F = None
            valid = ws[f"F{row}"].value

        raw_rows.append((sample_name, E, Y, U, F, valid))
        row += 1

    groups = {}
    for sample_name, E, Y, U, F, valid in raw_rows:
        base = get_subgroup(sample_name)
        groups.setdefault(base, []).append((sample_name, E, Y, U, F, valid))

    if include_fracture:
        headers = [
            "Sample Name",
            "Count",
            "E average",
            "E SDV",
            "E CV",
            "Yield average",
            "Yield SDV",
            "Yield CV",
            "US Average",
            "US SDV",
            "US CV",
            "FS Average",
            "FS SDV",
            "FS CV",
            "Flagged",
        ]
    else:
        headers = [
            "Sample Name",
            "Count",
            "E average",
            "E SDV",
            "E CV",
            "Yield average",
            "Yield SDV",
            "Yield CV",
            "US Average",
            "US SDV",
            "US CV",
            "Flagged",
        ]

    max_summary_cols = 15
    for r in range(1, ws.max_row + 1):
        for c in range(Summary_start_col, Summary_start_col + max_summary_cols):
            ws.cell(row=r, column=c).value = None

    for i, h in enumerate(headers, start=Summary_start_col):
        ws.cell(row=1, column=i, value=h)

    def clean_numeric_list(values):
        nums = []
        for v in values:
            try:
                x = float(v)
                if math.isfinite(x):
                    nums.append(x)
            except Exception:
                pass
        return nums

    def avg_or_none(vals):
        nums = clean_numeric_list(vals)
        return statistics.mean(nums) if nums else None

    def safe_std(vals):
        nums = clean_numeric_list(vals)
        return statistics.stdev(nums) if len(nums) >= 2 else None

    summary_row = 2

    for base_name, items in sorted(groups.items()):
        E_vals = [x[1] for x in items]
        Y_vals = [x[2] for x in items]
        U_vals = [x[3] for x in items]
        F_vals = [x[4] for x in items] if include_fracture else []

        invalid_count = sum(
            str(x[5]).strip().upper() == "NO"
            for x in items
            if x[5] is not None
        )

        e_avg = avg_or_none(E_vals)
        e_std = safe_std(E_vals)
        e_cv = safe_cv(E_vals)

        y_avg = avg_or_none(Y_vals)
        y_std = safe_std(Y_vals)
        y_cv = safe_cv(Y_vals)

        u_avg = avg_or_none(U_vals)
        u_std = safe_std(U_vals)
        u_cv = safe_cv(U_vals)

        flagged = False
        if e_cv is not None and e_cv > 0.10:
            flagged = True
        if y_cv is not None and y_cv > 0.10:
            flagged = True
        if u_cv is not None and u_cv > 0.10:
            flagged = True
        if invalid_count > 0:
            flagged = True

        ws.cell(row=summary_row, column=Summary_start_col + 0, value=base_name)
        ws.cell(row=summary_row, column=Summary_start_col + 1, value=len(items))

        ws.cell(row=summary_row, column=Summary_start_col + 2, value=e_avg)
        ws.cell(row=summary_row, column=Summary_start_col + 3, value=e_std)
        ws.cell(row=summary_row, column=Summary_start_col + 4, value=e_cv)

        ws.cell(row=summary_row, column=Summary_start_col + 5, value=y_avg)
        ws.cell(row=summary_row, column=Summary_start_col + 6, value=y_std)
        ws.cell(row=summary_row, column=Summary_start_col + 7, value=y_cv)

        ws.cell(row=summary_row, column=Summary_start_col + 8, value=u_avg)
        ws.cell(row=summary_row, column=Summary_start_col + 9, value=u_std)
        ws.cell(row=summary_row, column=Summary_start_col + 10, value=u_cv)

        if include_fracture:
            f_avg = avg_or_none(F_vals)
            f_std = safe_std(F_vals)
            f_cv = safe_cv(F_vals)

            ws.cell(row=summary_row, column=Summary_start_col + 11, value=f_avg)
            ws.cell(row=summary_row, column=Summary_start_col + 12, value=f_std)
            ws.cell(row=summary_row, column=Summary_start_col + 13, value=f_cv)
            ws.cell(row=summary_row, column=Summary_start_col + 14, value=flagged)
        else:
            ws.cell(row=summary_row, column=Summary_start_col + 11, value=flagged)

        summary_row += 1

    wb.save(filepath)


def highlight_flagged_samples(filepath, sheet_name=None):
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    yellow_fill = PatternFill(fill_type="solid", start_color="808000", end_color="808000")
    red_fill = PatternFill(fill_type="solid", start_color="FF5C5C", end_color="FF5C5C")
    no_fill = PatternFill(fill_type=None)

    def local_get_subgroup(name):
        if name is None:
            return ""
        return re.split(r"\s*-\s*\d+$", canonical_sample_name(name))[0]

    def is_bad_cv(val):
        if val is None or val == "":
            return False
        if isinstance(val, str):
            txt = val.strip().upper()
            if "#DIV/0!" in txt:
                return True
            if txt.endswith("%"):
                try:
                    return float(txt[:-1]) > 10
                except ValueError:
                    return False
        try:
            return float(val) > 0.10
        except Exception:
            return False

    has_fracture = str(ws.cell(row=1, column=Summary_start_col + 14).value).strip().upper() == "FLAGGED"

    if has_fracture:
        flagged_col = Summary_start_col + 14
        cv_cols = [Summary_start_col + 4, Summary_start_col + 7, Summary_start_col + 10, Summary_start_col + 13]
    else:
        flagged_col = Summary_start_col + 11
        cv_cols = [Summary_start_col + 4, Summary_start_col + 7, Summary_start_col + 10]

    flagged_samples = set()

    row = 2
    while ws.cell(row=row, column=Summary_start_col).value not in (None, ""):
        sample_name = ws.cell(row=row, column=Summary_start_col).value
        flagged = ws.cell(row=row, column=flagged_col).value

        if flagged is True or str(flagged).strip().upper() == "TRUE":
            flagged_samples.add(str(sample_name).strip())

        row += 1

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            cell.fill = no_fill

    row = 2
    while ws[f"A{row}"].value not in (None, ""):
        raw_name = str(ws[f"A{row}"].value).strip()
        base_name = local_get_subgroup(raw_name)
        if base_name in flagged_samples:
            ws[f"A{row}"].fill = yellow_fill
        row += 1

    row = 2
    while ws.cell(row=row, column=Summary_start_col).value not in (None, ""):
        for col in cv_cols:
            cell = ws.cell(row=row, column=col)
            if is_bad_cv(cell.value):
                cell.fill = red_fill
        row += 1

    wb.save(filepath)


def finalize_results_workbook(filepath, sheet_name=None):
    sort_results_workbook(filepath, sheet_name)

    name = Path(filepath).name.upper()
    include_fracture = "TENSION" in name

    build_summary_table(filepath, sheet_name, include_fracture=include_fracture)
    highlight_flagged_samples(filepath, sheet_name)
