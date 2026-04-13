from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
from petg_tension_core import render_petg_tension_graphs
from shared_core.adaptive_fit import (
    ElasticSampleAnalysis,
    NYLON_CANDIDATE_END_STRAIN_BANDS,
    canonical_sample_name,
    cleanup_stale_result_graphs,
    extract_elastic_candidates,
    format_sample_display_name,
    result_graph_variants,
    rerank_candidates_by_end_strain_bands,
)

from .constants import (
    EXPECTED_SAMPLE_SKIP_MARKERS,
    PETG_TENSION_FALLBACK_MIN_POINTS,
    TENSION_PRIMARY_R2_MIN,
    TENSION_SLOPE_CV_MAX,
    TENSION_STRESS_GATE_DELTA,
    TENSION_THRESHOLD_STRESS,
    TENSION_YIELD_METHOD_FIRST_MAX,
    TENSION_WINDOW_STEP,
    TENSION_YIELD_METHOD_LOCAL_PEAK,
    TENSION_YIELD_METHOD_PROOF,
    TENSION_ZERO_STRAIN_MAX,
)
from .io import clean_nylon_tension_curve, read_tension_xlsx, stitch_ext_lvdt_encoder
from .results import build_tension_result_row, get_subgroup, save_results_to_xlsx, write_tension_results
from .subgroup import resolve_subgroup_fit_decisions
from .yielding import (
    determine_tension_yield,
    find_tension_analysis_end_idx,
    has_tension_recovery_after_index,
)


_REJECTED_INVALID_CURVE = "rejected_invalid_curve"


@dataclass
class _TensionOutputArtifact:
    sample_name: str
    material: str
    is_valid: bool
    row: list
    graph_path: Path
    offset_strain: np.ndarray
    offset_stress: np.ndarray
    marker_strain: float
    marker_stress: float
    marker_label: str
    yield_method: str


_OFFSET_DISPLAY_PADDING_POINTS = 10
_WINDUP_DISPLAY_START_STRAIN_MIN = 0.01
_LINE_DISPLAY_BACKTRACK_RESIDUAL_FRACTION = 0.10
_LINE_DISPLAY_BACKTRACK_RESIDUAL_MIN = 0.5e6
_LINE_DISPLAY_BACKTRACK_MAX_STRAIN_DELTA = 0.003
_LINE_DISPLAY_BACKTRACK_MIN_START_STRESS_RATIO = 0.82


def _marker_label_for_tension_point(material: str, yield_method: str) -> str:
    if yield_method in (TENSION_YIELD_METHOD_LOCAL_PEAK, TENSION_YIELD_METHOD_FIRST_MAX):
        return "Peak Point"
    return "Yield Point"


def _format_offset_curve_for_marker_display(
    offset_strain,
    offset_stress,
    marker_strain,
    marker_stress,
    marker_label: str,
    *,
    min_display_stress: float | None = None,
    min_display_strain: float | None = None,
):
    offset_strain = np.asarray(offset_strain, dtype=float)
    offset_stress = np.asarray(offset_stress, dtype=float)

    if (
        offset_strain.size == 0
        or offset_stress.size == 0
        or not np.isfinite(marker_strain)
        or not np.isfinite(marker_stress)
    ):
        return offset_strain, offset_stress

    display_start = 0
    if min_display_stress is not None and np.isfinite(min_display_stress):
        visible_start = next(
            (index for index, value in enumerate(offset_stress) if value >= float(min_display_stress)),
            None,
        )
        if visible_start is not None:
            display_start = max(visible_start - _OFFSET_DISPLAY_PADDING_POINTS, 0)
    if min_display_strain is not None and np.isfinite(min_display_strain):
        visible_start = next(
            (index for index, value in enumerate(offset_strain) if value >= float(min_display_strain)),
            None,
        )
        if visible_start is not None:
            display_start = max(display_start, visible_start)

    if marker_label == "Fracture Point":
        return offset_strain[display_start:], offset_stress[display_start:]

    crossing_index = next(
        (index for index, value in enumerate(offset_strain) if value >= float(marker_strain)),
        None,
    )
    if crossing_index is None:
        return offset_strain[display_start:], offset_stress[display_start:]

    display_end = min(crossing_index + _OFFSET_DISPLAY_PADDING_POINTS, offset_strain.size)
    if display_end <= display_start:
        return offset_strain, offset_stress
    return offset_strain[display_start:display_end], offset_stress[display_start:display_end]


def _finite_plot_bounds(*series, fallback: tuple[float, float]):
    finite_values = []
    for values in series:
        array = np.asarray(values, dtype=float).reshape(-1)
        if array.size == 0:
            continue
        finite = array[np.isfinite(array)]
        if finite.size > 0:
            finite_values.append(finite)

    if not finite_values:
        return fallback

    combined = np.concatenate(finite_values)
    return float(np.min(combined)), float(np.max(combined))


def _should_hide_offset_curve(
    material: str,
    yield_method: str,
    yield_strain: float,
    yield_stress: float,
) -> bool:
    material_upper = material.upper()
    if material_upper in {"NYLON", "PETG"}:
        return False
    return not (np.isfinite(yield_strain) and np.isfinite(yield_stress))


def _tension_display_start_idx(strain, stress) -> int:
    strain = np.asarray(strain, dtype=float)
    stress = np.asarray(stress, dtype=float)
    if strain.size == 0 or stress.size == 0:
        return 0

    threshold_hits = np.flatnonzero(np.isfinite(stress) & (stress >= float(TENSION_THRESHOLD_STRESS)))
    if threshold_hits.size == 0:
        return 0

    first_loaded_idx = int(threshold_hits[0])
    if first_loaded_idx <= _OFFSET_DISPLAY_PADDING_POINTS:
        return 0

    loaded_strain = float(strain[min(first_loaded_idx, strain.size - 1)])
    if not np.isfinite(loaded_strain) or loaded_strain < _WINDUP_DISPLAY_START_STRAIN_MIN:
        return 0

    return max(first_loaded_idx - _OFFSET_DISPLAY_PADDING_POINTS, 0)


def _line_display_start_idx(strain, stress, start_idx, end_idx, slope, intercept) -> int:
    strain = np.asarray(strain, dtype=float)
    stress = np.asarray(stress, dtype=float)
    if strain.size == 0 or stress.size == 0:
        return 0
    if start_idx is None or end_idx is None:
        return 0
    if not np.isfinite(slope) or not np.isfinite(intercept):
        return max(int(start_idx), 0)

    start = max(int(start_idx), 0)
    end = min(int(end_idx), len(stress) - 1)
    if end < start:
        return start

    fit_stress = stress[start : end + 1]
    finite_fit_stress = fit_stress[np.isfinite(fit_stress)]
    if finite_fit_stress.size == 0:
        return start

    fit_start_stress = float(stress[start])
    min_start_stress = fit_start_stress * _LINE_DISPLAY_BACKTRACK_MIN_START_STRESS_RATIO
    min_display_strain = float(strain[start]) - _LINE_DISPLAY_BACKTRACK_MAX_STRAIN_DELTA
    fit_stress_span = float(np.max(finite_fit_stress) - np.min(finite_fit_stress))
    residual_tol = max(
        float(fit_stress_span) * _LINE_DISPLAY_BACKTRACK_RESIDUAL_FRACTION,
        _LINE_DISPLAY_BACKTRACK_RESIDUAL_MIN,
    )

    display_start = start
    while display_start > 0:
        prev_idx = display_start - 1
        if float(strain[prev_idx]) < min_display_strain:
            break
        predicted_stress = float(slope) * float(strain[prev_idx]) + float(intercept)
        actual_stress = float(stress[prev_idx])
        if not np.isfinite(predicted_stress) or not np.isfinite(actual_stress):
            break
        if np.isfinite(fit_start_stress) and actual_stress < min_start_stress:
            break
        if abs(actual_stress - predicted_stress) > residual_tol:
            break
        display_start = prev_idx

    return display_start


def _has_invalid_pre_failure_negative_stress(stress, *, threshold_stress: float) -> bool:
    stress = np.asarray(stress, dtype=float)
    analysis_end_idx = find_tension_analysis_end_idx(stress)
    if stress.size == 0 or analysis_end_idx is None:
        return False

    threshold_hits = np.flatnonzero(stress >= float(threshold_stress))
    if threshold_hits.size == 0 or analysis_end_idx < int(threshold_hits[0]):
        return False

    analysis_slice = stress[int(threshold_hits[0]) : analysis_end_idx + 1]
    negative_mask = analysis_slice < 0.0
    if not np.any(negative_mask):
        return False

    segment_starts = np.flatnonzero(
        negative_mask & np.concatenate((np.array([True]), ~negative_mask[:-1]))
    )
    for rel_start in segment_starts:
        segment_start = int(threshold_hits[0]) + int(rel_start)
        pre_segment = stress[int(threshold_hits[0]) : segment_start]
        finite_pre_segment = pre_segment[np.isfinite(pre_segment)]
        reference_stress = float(np.nanmax(finite_pre_segment)) if finite_pre_segment.size > 0 else float(threshold_stress)
        if has_tension_recovery_after_index(
            stress[: analysis_end_idx + 1],
            segment_start + 1,
            reference_stress=reference_stress,
        ):
            continue
        return True

    return False


def _apply_tension_quality_rejection(sample: ElasticSampleAnalysis):
    if not _has_invalid_pre_failure_negative_stress(
        sample.stress,
        threshold_stress=TENSION_THRESHOLD_STRESS,
    ):
        return sample

    sample.selected_candidate = None
    sample.is_valid = False
    sample.fit_mode = ""
    sample.decision_reason = _REJECTED_INVALID_CURVE
    return sample


def _trim_tension_curve_to_analysis_segment(strain, stress):
    strain = np.asarray(strain, dtype=float)
    stress = np.asarray(stress, dtype=float)
    analysis_end_idx = find_tension_analysis_end_idx(stress)
    if analysis_end_idx is None or strain.size == 0 or stress.size == 0:
        return strain, stress
    trimmed_end_idx = min(int(analysis_end_idx), len(strain) - 1, len(stress) - 1)
    return strain[: trimmed_end_idx + 1], stress[: trimmed_end_idx + 1]


def _curve_peak_strain(strain, stress):
    strain = np.asarray(strain, dtype=float)
    stress = np.asarray(stress, dtype=float)
    if strain.size == 0 or stress.size == 0:
        return np.nan
    peak_idx = int(np.nanargmax(stress))
    return float(strain[peak_idx])


def _nylon_candidate_band_index(candidate, peak_strain: float) -> int:
    if not np.isfinite(peak_strain) or peak_strain <= 0.0:
        return len(NYLON_CANDIDATE_END_STRAIN_BANDS)
    end_ratio = float(candidate.end_strain) / float(peak_strain)
    for idx, limit in enumerate(NYLON_CANDIDATE_END_STRAIN_BANDS):
        if end_ratio <= float(limit):
            return idx
    return len(NYLON_CANDIDATE_END_STRAIN_BANDS)


def _candidate_supports_nylon_proof_yield(strain, stress, candidate) -> bool:
    (
        _,
        _,
        _,
        _,
        _,
        _,
        yield_method,
    ) = determine_tension_yield(
        strain,
        stress,
        candidate.slope,
        0.002,
        candidate.start_idx,
        candidate.end_idx,
        candidate.intercept,
        material="NYLON",
    )
    return yield_method == TENSION_YIELD_METHOD_PROOF


def _prioritize_nylon_candidates_for_yield(strain, stress, candidates):
    candidate_list = list(candidates)
    if not candidate_list:
        return []

    peak_strain = _curve_peak_strain(strain, stress)
    banded_candidates = rerank_candidates_by_end_strain_bands(
        candidate_list,
        peak_strain=peak_strain,
    )

    proof_candidates = []
    proof_candidate_ids = set()
    for band_index in range(len(NYLON_CANDIDATE_END_STRAIN_BANDS) + 1):
        band_candidates = [
            candidate
            for candidate in banded_candidates
            if _nylon_candidate_band_index(candidate, peak_strain) == band_index
        ]
        band_proof_candidates = [
            candidate
            for candidate in band_candidates
            if _candidate_supports_nylon_proof_yield(strain, stress, candidate)
        ]
        if band_proof_candidates:
            proof_candidates = band_proof_candidates
            proof_candidate_ids = {id(candidate) for candidate in band_proof_candidates}
            break

    if not proof_candidates:
        return banded_candidates

    remaining_candidates = [
        candidate
        for candidate in banded_candidates
        if id(candidate) not in proof_candidate_ids
    ]
    return proof_candidates + remaining_candidates


def _is_present(value) -> bool:
    return value is not None and not (isinstance(value, (float, np.floating)) and np.isnan(value))

def _load_saved_result_rows(excel_path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, data_only=False)
    worksheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    rows = {}
    for row_idx in range(2, worksheet.max_row + 1):
        values = {
            headers[column - 1]: worksheet.cell(row=row_idx, column=column).value
            for column in range(1, len(headers) + 1)
        }
        sample_name = canonical_sample_name(values.get("Sample Name"))
        if sample_name:
            rows[sample_name] = values
    workbook.close()
    return rows


def _audit_nylon_outputs(artifacts, excel_path: Path):
    if not artifacts:
        return

    saved_rows = _load_saved_result_rows(excel_path)
    for artifact in artifacts:
        if artifact.material.upper() != "NYLON":
            continue

        canonical_name = canonical_sample_name(artifact.sample_name)
        if canonical_name not in saved_rows:
            raise ValueError(f"Missing saved nylon result row for '{canonical_name}'")
        if not artifact.graph_path.exists():
            raise ValueError(f"Missing nylon graph output for '{canonical_name}'")

        remaining_variants = [
            path
            for path in result_graph_variants(artifact.graph_path.parent, artifact.sample_name)
            if path.exists()
        ]
        if remaining_variants != [artifact.graph_path]:
            raise ValueError(f"Stale nylon graph variants remain for '{canonical_name}'")

        if not artifact.is_valid:
            continue

        row = saved_rows[canonical_name]
        graph_html = artifact.graph_path.read_text(encoding="utf-8", errors="ignore")

        if artifact.offset_strain.size == 0 or artifact.offset_stress.size == 0:
            raise ValueError(f"Nylon graph for '{canonical_name}' is missing offset data")
        if "Offset Curve" not in graph_html:
            raise ValueError(f"Nylon graph for '{canonical_name}' is missing the offset trace")

        if artifact.yield_method == TENSION_YIELD_METHOD_PROOF:
            if not _is_present(row.get("Yield Strength (MPa)")):
                raise ValueError(f"Nylon proof result for '{canonical_name}' is missing yield strength")
            if "Yield Point" not in graph_html or "Peak Point" in graph_html:
                raise ValueError(f"Nylon proof graph for '{canonical_name}' has incorrect marker labels")
            expected_stress = float(
                np.interp(float(artifact.marker_strain), artifact.offset_strain, artifact.offset_stress)
            )
            if not np.isfinite(artifact.marker_stress) or not np.isclose(
                float(artifact.marker_stress),
                expected_stress,
                atol=1.0e3,
            ):
                raise ValueError(f"Nylon proof marker for '{canonical_name}' is not on the offset curve")
        elif artifact.yield_method == TENSION_YIELD_METHOD_LOCAL_PEAK:
            if _is_present(row.get("Yield Strength (MPa)")):
                raise ValueError(f"Nylon peak-only result for '{canonical_name}' should have blank yield strength")
            if not _is_present(row.get("Ultimate Strength (MPa)")):
                raise ValueError(f"Nylon peak-only result for '{canonical_name}' is missing ultimate strength")
            if "Peak Point" not in graph_html or "Yield Point" in graph_html:
                raise ValueError(f"Nylon peak-only graph for '{canonical_name}' has incorrect marker labels")
        else:
            raise ValueError(f"Nylon graph for '{canonical_name}' has unsupported yield method '{artifact.yield_method}'")


def select_tension_fit(
    strain,
    stress,
    *,
    window_min=75,
    window_step=TENSION_WINDOW_STEP,
    r2_min=TENSION_PRIMARY_R2_MIN,
    threshold_stress=TENSION_THRESHOLD_STRESS,
    stress_gate_delta=TENSION_STRESS_GATE_DELTA,
    zero_strain_max=TENSION_ZERO_STRAIN_MAX,
    slope_cv_max=TENSION_SLOPE_CV_MAX,
    fallback_min_points=20,
):
    return extract_elastic_candidates(
        strain,
        stress,
        window_min=window_min,
        window_step=window_step,
        r2_min=r2_min,
        threshold_stress=threshold_stress,
        stress_gate_delta=stress_gate_delta,
        zero_strain_max=zero_strain_max,
        slope_cv_max=slope_cv_max,
        fallback_min_points=fallback_min_points,
    )


def find_linear_region_properties_V3(
    strain,
    stress,
    *,
    window_min=75,
    window_step=TENSION_WINDOW_STEP,
    r2_min=TENSION_PRIMARY_R2_MIN,
    threshold_stress=TENSION_THRESHOLD_STRESS,
    fallback_min_points=20,
):
    _ = fallback_min_points
    search = select_tension_fit(
        strain,
        stress,
        window_min=window_min,
        window_step=window_step,
        r2_min=r2_min,
        threshold_stress=threshold_stress,
        fallback_min_points=fallback_min_points,
    )

    if not search.candidates:
        return np.nan, None, None, np.nan, search.strain, search.stress, np.nan

    best = search.candidates[0]
    return (
        best.slope,
        best.start_idx,
        best.end_idx,
        best.r2,
        search.strain,
        search.stress,
        best.intercept,
    )


def offset_intersection(
    strain_clean,
    stress_clean,
    E,
    offset,
    linear_start,
    linear_end,
    intercept,
    *,
    material="",
):
    (
        offset_strain,
        offset_stress,
        yield_strain,
        yield_stress,
        ultimate_strength,
        fracture_strength,
        yield_method,
    ) = determine_tension_yield(
        strain_clean,
        stress_clean,
        E,
        offset,
        linear_start,
        linear_end,
        intercept,
        material=material,
    )
    if _should_hide_offset_curve(material, yield_method, yield_strain, yield_stress):
        offset_strain = np.array([], dtype=float)
        offset_stress = np.array([], dtype=float)
    return offset_strain, offset_stress, yield_strain, yield_stress, ultimate_strength, fracture_strength


def plot_stress_strain(strain, stress, start_idx, end_idx, slope, offset_strain, offset_stress, offset_intersection_strain, offset_intersection_stress, intercept):
    import matplotlib.pyplot as plt

    strain = np.asarray(strain, dtype=float)
    stress = np.asarray(stress, dtype=float)
    display_start_idx = _tension_display_start_idx(strain, stress)
    if display_start_idx > 0:
        strain = strain[display_start_idx:]
        stress = stress[display_start_idx:]
        if start_idx is not None:
            start_idx = max(int(start_idx) - display_start_idx, 0)
        if end_idx is not None:
            end_idx = max(int(end_idx) - display_start_idx, -1)

    line_start_idx = _line_display_start_idx(strain, stress, start_idx, end_idx, slope, intercept)
    offset_min_display_strain = float(strain[line_start_idx]) if len(strain) > 0 else None
    offset_strain, offset_stress = _format_offset_curve_for_marker_display(
        offset_strain,
        offset_stress,
        offset_intersection_strain,
        offset_intersection_stress,
        "Yield Point",
        min_display_stress=float(np.nanmin(np.asarray(stress, dtype=float))) if len(stress) > 0 else None,
        min_display_strain=offset_min_display_strain,
    )

    plt.figure(figsize=(8, 6))
    plt.plot(strain, stress, label="Stress-Strain Curve")
    plt.plot(offset_strain, offset_stress, label="Offset Stress-Strain Curve")
    plt.scatter(offset_intersection_strain, offset_intersection_stress, color="red", label="Offset Intersection")
    plt.plot(strain[line_start_idx:end_idx + 1], stress[line_start_idx:end_idx + 1], linewidth=3, label="Linear Region")
    x_lin = strain[line_start_idx:end_idx + 1]
    y_lin = slope * x_lin + intercept
    plt.plot(x_lin, y_lin, linestyle="--", label="Linear Fit")
    plt.xlabel("Strain")
    plt.ylabel("Stress")
    plt.title("Stress-Strain Curve")
    plt.grid(True)
    plt.legend()
    plt.show()


def plot_stress_strain_interactive(
    strain,
    stress,
    start_idx,
    end_idx,
    slope,
    intercept,
    offset_strain,
    offset_stress,
    yield_strain,
    yield_stress,
    yield_label,
    save_path,
    file_name,
):
    import plotly.graph_objects as go

    strain = np.asarray(strain, dtype=float)
    stress = np.asarray(stress, dtype=float)
    offset_strain = np.asarray(offset_strain, dtype=float)
    offset_stress = np.asarray(offset_stress, dtype=float)
    analysis_end_idx = find_tension_analysis_end_idx(stress)
    if analysis_end_idx is not None and strain.size > 0 and stress.size > 0:
        analysis_end_idx = min(int(analysis_end_idx), len(strain) - 1, len(stress) - 1)
        strain = strain[: analysis_end_idx + 1]
        stress = stress[: analysis_end_idx + 1]
        if offset_strain.size > 0 and offset_stress.size > 0:
            offset_mask = offset_strain <= float(strain[-1])
            offset_strain = offset_strain[offset_mask]
            offset_stress = offset_stress[offset_mask]

    display_start_idx = _tension_display_start_idx(strain, stress)
    if display_start_idx > 0:
        strain = strain[display_start_idx:]
        stress = stress[display_start_idx:]
        if start_idx is not None:
            start_idx = max(int(start_idx) - display_start_idx, 0)
        if end_idx is not None:
            end_idx = max(int(end_idx) - display_start_idx, -1)

    line_start_idx = _line_display_start_idx(strain, stress, start_idx, end_idx, slope, intercept)
    offset_min_display_strain = float(strain[line_start_idx]) if strain.size > 0 else None
    stress = stress / 1e6
    offset_stress = offset_stress / 1e6
    yield_stress = yield_stress / 1e6
    slope = slope / 1e6
    intercept = intercept / 1e6

    offset_strain, offset_stress = _format_offset_curve_for_marker_display(
        offset_strain,
        offset_stress,
        yield_strain,
        yield_stress,
        yield_label,
        min_display_stress=float(np.nanmin(stress)) if stress.size > 0 else None,
        min_display_strain=offset_min_display_strain,
    )

    has_linear_region = (
        start_idx is not None
        and end_idx is not None
        and 0 <= start_idx <= end_idx < len(strain)
        and np.isfinite(slope)
        and np.isfinite(intercept)
    )

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=strain, y=stress, mode="lines", name="Stress-Strain Curve"))
    x_series = [strain]
    y_series = [stress]

    if has_linear_region:
        fig.add_trace(go.Scatter(
            x=strain[line_start_idx:end_idx + 1],
            y=stress[line_start_idx:end_idx + 1],
            mode="lines",
            line=dict(width=6),
            name="Linear Region",
        ))
        x_lin = strain[line_start_idx:end_idx + 1]
        y_lin = slope * x_lin + intercept
        x_series.append(x_lin)
        y_series.append(y_lin)
        fig.add_trace(go.Scatter(
            x=x_lin,
            y=y_lin,
            mode="lines",
            line=dict(dash="dash"),
            name="Linear Fit",
        ))

    if len(offset_strain) > 0 and len(offset_stress) > 0:
        x_series.append(offset_strain)
        y_series.append(offset_stress)
        fig.add_trace(go.Scatter(x=offset_strain, y=offset_stress, mode="lines", name="Offset Curve"))

    if np.isfinite(yield_strain) and np.isfinite(yield_stress):
        x_series.append(np.array([yield_strain], dtype=float))
        y_series.append(np.array([yield_stress], dtype=float))
        fig.add_trace(go.Scatter(
            x=[yield_strain],
            y=[yield_stress],
            mode="markers",
            marker=dict(size=10, color="red"),
            name=yield_label,
        ))

    x_min, x_max = _finite_plot_bounds(*x_series, fallback=(0.0, 1.0))
    y_min, y_max = _finite_plot_bounds(*y_series, fallback=(0.0, 1.0))
    x_pad = 0.05 * max(x_max - x_min, 1.0e-6)
    y_pad = 0.05 * max(y_max - y_min, 1.0e-6)

    fig.update_layout(
        title="Stress-Strain Curve of " + file_name,
        title_x=0.5,
        xaxis_title="Strain",
        yaxis_title="Stress (MPa)",
        template="plotly_white",
        xaxis=dict(range=[x_min - x_pad, x_max + x_pad]),
        yaxis=dict(range=[y_min - y_pad, y_max + y_pad]),
    )

    save_path.parent.mkdir(parents=True, exist_ok=True)
    fig.write_html(save_path)
    return save_path


def _is_expected_sample_skip(error: Exception) -> bool:
    message = str(error)
    return message.startswith("Sample '") and any(marker in message for marker in EXPECTED_SAMPLE_SKIP_MARKERS)


def _material_name_from_data_path(path: Path) -> str:
    try:
        data_idx = path.parts.index("DATA")
    except ValueError:
        return path.parent.name or path.name
    material_idx = data_idx + 1
    if material_idx < len(path.parts):
        return path.parts[material_idx]
    return path.parent.name or path.name


def _analyze_tension_file(file_name: Path):
    encoder_strain, lvdt_strain, extensometer, _, stress_Pa = read_tension_xlsx(
        file_name,
        prompt_for_area=True,
    )
    material = file_name.parts[file_name.parts.index("DATA") + 1]

    strain, _ = stitch_ext_lvdt_encoder(
        extensometer,
        lvdt_strain,
        encoder_strain,
        tol_strain=0.001,
        window=15,
        min_start=15,
    )
    if material.upper() == "NYLON":
        strain, stress_Pa = clean_nylon_tension_curve(strain, stress_Pa)
        strain, stress_Pa = _trim_tension_curve_to_analysis_segment(strain, stress_Pa)

    window_min = 100 if "100" in file_name.stem else 75
    search = select_tension_fit(
        strain,
        stress_Pa,
        window_min=window_min,
        window_step=TENSION_WINDOW_STEP,
        r2_min=TENSION_PRIMARY_R2_MIN,
        threshold_stress=TENSION_THRESHOLD_STRESS,
        stress_gate_delta=TENSION_STRESS_GATE_DELTA,
        zero_strain_max=TENSION_ZERO_STRAIN_MAX,
        slope_cv_max=TENSION_SLOPE_CV_MAX,
        fallback_min_points=PETG_TENSION_FALLBACK_MIN_POINTS if material.upper() == "PETG" else 20,
    )
    candidates = list(search.candidates)
    if material.upper() == "NYLON":
        candidates = _prioritize_nylon_candidates_for_yield(
            search.strain,
            search.stress,
            candidates,
        )

    return ElasticSampleAnalysis(
        sample_name=file_name.stem,
        file_path=file_name,
        material=material,
        strain=np.asarray(search.strain, dtype=float),
        stress=np.asarray(search.stress, dtype=float),
        candidates=candidates,
    )


def _finalize_tension_metrics(
    sample: ElasticSampleAnalysis,
    graph_root: Path,
    *,
    petg_component_graphs: bool = True,
):
    _apply_tension_quality_rejection(sample)
    candidate = sample.selected_candidate
    graph_name = format_sample_display_name(sample.sample_name, sample.is_valid, sample.decision_reason)

    if sample.is_valid and candidate is not None:
        (
            offset_strain,
            offset_stress,
            yield_strain,
            yield_stress,
            ultimate_strength,
            fracture_strength,
            yield_method,
        ) = determine_tension_yield(
            sample.strain,
            sample.stress,
            candidate.slope,
            0.002,
            candidate.start_idx,
            candidate.end_idx,
            candidate.intercept,
            material=sample.material,
        )
        E = candidate.slope
        start_idx = candidate.start_idx
        end_idx = candidate.end_idx
        intercept = candidate.intercept
        fit_start_strain = candidate.start_strain
        fit_end_strain = candidate.end_strain
        marker_strain = yield_strain
        marker_stress = yield_stress
        marker_label = _marker_label_for_tension_point(sample.material, yield_method)
        stored_yield_stress = yield_stress if yield_method == TENSION_YIELD_METHOD_PROOF else np.nan
        if _should_hide_offset_curve(sample.material, yield_method, yield_strain, yield_stress):
            offset_strain = np.array([], dtype=float)
            offset_stress = np.array([], dtype=float)
    else:
        offset_strain = np.array([], dtype=float)
        offset_stress = np.array([], dtype=float)
        marker_strain = np.nan
        marker_stress = np.nan
        ultimate_strength = float(np.nanmax(sample.stress)) if sample.stress.size > 0 else np.nan
        finite_idx = np.where(np.isfinite(sample.stress))[0]
        fracture_strength = float(sample.stress[finite_idx[-1]]) if finite_idx.size > 0 else np.nan
        E = np.nan
        start_idx = None
        end_idx = None
        intercept = np.nan
        fit_start_strain = np.nan
        fit_end_strain = np.nan
        marker_label = "Yield Point"
        yield_method = ""
        stored_yield_stress = np.nan

    graph_path = graph_root / sample.material / "TENSION" / f"{graph_name}.html"
    cleanup_stale_result_graphs(graph_path, sample.sample_name)
    if sample.material.upper() == "PETG":
        graph_bundle = render_petg_tension_graphs(
            sample_name=sample.sample_name,
            display_name=graph_name,
            graph_path=graph_path,
            strain=sample.strain,
            stress=sample.stress,
            start_idx=start_idx,
            end_idx=end_idx,
            slope=E,
            intercept=intercept,
            offset_strain=offset_strain,
            offset_stress=offset_stress,
            yield_strain=marker_strain,
            yield_stress=marker_stress,
            yield_method=yield_method,
            write_component_graphs=petg_component_graphs,
        )
        graph_link = graph_bundle.primary_graph_path
    else:
        graph_link = plot_stress_strain_interactive(
            sample.strain,
            sample.stress,
            start_idx,
            end_idx,
            E,
            intercept,
            offset_strain,
            offset_stress,
            marker_strain,
            marker_stress,
            marker_label,
            save_path=graph_path,
            file_name=graph_name,
        )

    row = build_tension_result_row(
        sample_name=sample.sample_name,
        E=E,
        yield_strength=stored_yield_stress,
        ultimate_strength=ultimate_strength,
        fracture_strength=fracture_strength,
        plot_path=graph_link,
        is_valid=sample.is_valid,
        fit_start_strain=fit_start_strain,
        fit_end_strain=fit_end_strain,
        fit_mode=sample.fit_mode,
        decision_reason=sample.decision_reason,
    )

    return _TensionOutputArtifact(
        sample_name=sample.sample_name,
        material=sample.material,
        is_valid=sample.is_valid,
        row=row,
        graph_path=graph_link,
        offset_strain=np.asarray(offset_strain, dtype=float),
        offset_stress=np.asarray(offset_stress, dtype=float),
        marker_strain=float(marker_strain) if np.isfinite(marker_strain) else np.nan,
        marker_stress=float(marker_stress) if np.isfinite(marker_stress) else np.nan,
        marker_label=marker_label,
        yield_method=yield_method,
    )


def process_tension_file(file_name: Path, *, petg_component_graphs: bool = True):
    here = Path(__file__).resolve().parent.parent
    sample = _analyze_tension_file(file_name)
    resolve_subgroup_fit_decisions([sample])
    _apply_tension_quality_rejection(sample)

    excel_path = here / "STATS" / sample.material / f"{sample.material} TENSION RESULTS.xlsx"
    artifact = _finalize_tension_metrics(
        sample,
        here / "GRAPHS",
        petg_component_graphs=petg_component_graphs,
    )
    row = artifact.row
    save_results_to_xlsx(
        sample_name=row[0],
        E=np.nan if row[1] is None or not np.isfinite(row[1]) else float(row[1]) * 1e9,
        yield_strength=np.nan if row[2] is None or not np.isfinite(row[2]) else float(row[2]) * 1e6,
        ultimate_strength=np.nan if row[3] is None or not np.isfinite(row[3]) else float(row[3]) * 1e6,
        fracture_strength=np.nan if row[4] is None or not np.isfinite(row[4]) else float(row[4]) * 1e6,
        output_path=excel_path,
        plot_path=artifact.graph_path,
        is_valid=row[6] == "Yes",
        fit_start_strain=row[7],
        fit_end_strain=row[8],
        fit_mode=row[9],
        decision_reason=row[10],
    )
    if sample.material.upper() == "NYLON":
        _audit_nylon_outputs([artifact], excel_path)
    return sample


def process_tension_folder(folder: Path, *, petg_component_graphs: bool = True):
    here = Path(__file__).resolve().parent.parent
    material_hint = _material_name_from_data_path(folder)
    files = sorted(folder.glob("*.xlsx"))
    grouped_samples: dict[str, list[ElasticSampleAnalysis]] = {}
    skipped_count = 0
    failed_count = 0

    print(
        f"Tension scan starting: {material_hint} ({len(files)} workbook(s) in {folder.name})"
    )

    for file in files:
        try:
            sample = _analyze_tension_file(file)
        except Exception as e:
            label = "Skipped" if _is_expected_sample_skip(e) else "Failed"
            print(f"{label}: {file.name} -> {e}")
            if label == "Skipped":
                skipped_count += 1
            else:
                failed_count += 1
            continue
        subgroup = get_subgroup(sample.sample_name)
        grouped_samples.setdefault(subgroup, []).append(sample)

    all_samples = []
    for subgroup in sorted(grouped_samples):
        subgroup_samples = grouped_samples[subgroup]
        resolve_subgroup_fit_decisions(subgroup_samples)
        for sample in subgroup_samples:
            _apply_tension_quality_rejection(sample)
        all_samples.extend(subgroup_samples)

    print(
        "Tension scan complete: "
        f"{len(all_samples)} analyzable, {skipped_count} skipped, {failed_count} failed."
    )

    if not all_samples:
        print(f"Tension processing finished with no analyzable samples for {material_hint}.")
        return []

    print(
        "Tension output starting: "
        f"{material_hint} ({len(all_samples)} sample(s)); "
        f"PETG component graphs {'on' if petg_component_graphs else 'off'}."
    )
    artifacts = []
    total_samples = len(all_samples)
    for index, sample in enumerate(all_samples, start=1):
        artifact = _finalize_tension_metrics(
            sample,
            here / "GRAPHS",
            petg_component_graphs=petg_component_graphs,
        )
        artifacts.append(artifact)
        print(f"Tension output {index}/{total_samples}: {sample.sample_name}")

    rows = [artifact.row for artifact in artifacts]
    material = all_samples[0].material
    excel_path = here / "STATS" / material / f"{material} TENSION RESULTS.xlsx"
    write_tension_results(excel_path, rows)
    print(f"Tension results saved: {excel_path.name}")
    if material.upper() == "NYLON":
        _audit_nylon_outputs(artifacts, excel_path)
        print("Tension audit complete: NYLON")
    return all_samples
